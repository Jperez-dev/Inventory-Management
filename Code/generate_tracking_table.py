#!/usr/bin/python3

import openpyxl
from openpyxl.styles import Alignment
import sys
import os
from datetime import datetime

def update_tracking_history(withdrawal_control_number, requestor_name, withdrawal_date, pdf_filename, approval_manager, approval_spare_parts):
    # Path to the history file
    tracking_folder = "C:/Users/admin/git-directories/Data_Management/Inventory_Management/Withdrawal_Tracking/"
    history_file = tracking_folder + "Withdrawal_History_" + datetime.now().strftime("%Y-%m") + ".xlsx"
    
    # Check if the history file exists, create it if not
    if not os.path.exists(history_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Create headers (excluding manager and spare parts approvals)
        ws.append(["Withdrawal Control Number", "Requestor's Name", "Date Requested", "PDF Filename", "Status"])
    else:
        wb = openpyxl.load_workbook(history_file)
        ws = wb.active
    
    # Determine status (Approved or Rejected based on manager and spare parts approvals)
    if approval_manager == "Approved" and approval_spare_parts == "Approved":
        status = "Approved"
    else:
        status = "Rejected"

    # Create a clickable hyperlink to the PDF
    pdf_link = f'=HYPERLINK("C:/Users/admin/git-directories/Data_Management/Inventory_Management/Withdrawals/{pdf_filename}", "{pdf_filename}")'

    # Append new row with withdrawal details
    ws.append([withdrawal_control_number, requestor_name, withdrawal_date, pdf_link, status])

    # Get the row number of the last row that was added
    last_row = ws.max_row

    # Loop through each cell in the last row and set the alignment to center
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=last_row, column=col)
        
        # Stop when an empty cell is found
        if cell.value is None:
            break
        
        # Apply alignment to non-empty cells
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the updated tracking file
    wb.save(history_file)

# Receive arguments from VBA
if __name__ == "__main__":
    update_tracking_history(*sys.argv[1:])

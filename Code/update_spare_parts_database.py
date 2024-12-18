#!/usr/bin/python3

import openpyxl
import sys

def update_database(withdrawal_path):

    # Paths to your files
    withdrawal_form_path = withdrawal_path.replace('\\', '/')
    db_path = 'C:/Users/admin/git-directories/Data_Management/Inventory_Management/Spare_Parts_Database.xlsx'

    # Load the withdrawal form and the spare parts database
    wb_withdrawal = openpyxl.load_workbook(withdrawal_form_path)
    ws_withdrawal = wb_withdrawal['Finalized_Withdrawal_Form']  # Sheet for withdrawal form

    wb_db = openpyxl.load_workbook(db_path)
    ws_db = wb_db['Spare_Parts']  # Sheet for spare parts database

    # Loop through the item codes in the withdrawal form (C10:C19)
    for row in range(10, 20):
        item_code = ws_withdrawal[f'C{row}'].value  # Item code in C10:C19
        order_qty = ws_withdrawal[f'Q{row}'].value  # Issued Qty Column

        # Check if item code and order qty are non-empty
        if item_code and order_qty:  # Only proceed if both item_code and order_qty are non-empty
            # Find the corresponding available balance in the spare parts database
            for db_row in range(2, ws_db.max_row + 1): 
                db_item_code = ws_db[f'A{db_row}'].value  # Item code columns
                if db_item_code == item_code:
                    available_balance = ws_db[f'G{db_row}'].value  # Available Balance columns

                    # Calculate the new balance
                    if available_balance is not None:
                        new_balance = available_balance - order_qty
                        # Update the available balance in the spare parts database
                        ws_db[f'G{db_row}'].value = new_balance

                    break  # Exit once the correct item code is found (this ensures we don’t check more rows for the same item code)

    # Save the updated spare parts database
    wb_db.save(db_path)

if __name__ == "__main__":
    update_database(sys.argv[1])

